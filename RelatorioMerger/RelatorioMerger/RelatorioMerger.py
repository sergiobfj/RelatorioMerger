#Instalações e Importações
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import customtkinter as ctk
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

#Caminho dos Arquivos
path_posVenda = ''
path_vendasI = ''

# -=-=-=-=- PANDAS -=-=-=-=-=-

def mesclarVendasI(caminho_posVenda, caminho_vendasI):
    caminho = os.path.dirname(path_vendasI)

    #Leitura dos arquivos
    posVenda_df = pd.read_excel(caminho_posVenda)
    vendasI_df = pd.read_excel(caminho_vendasI)

    #Merge de dados
    vendasI_df = vendasI_df.merge(posVenda_df, on='Título')

    #Tratamento de dados
    colunas_drop = [
        'Código da Proposta_x','Cód. Cliente do Cliente','Cliente_x',
        'Responsável','Valor','Estágio','Dias no estágio','Cidade do Cliente',
        'Endereço do Cliente','Potência do Módulo do Cliente','Marca do Inversor do Cliente',
        'Potência do Inversor do Cliente','Qtd. Inversores do Cliente','Qtd. Módulo do Cliente'
    ]
    vendasI_df = vendasI_df.drop(colunas_drop, axis=1, errors="ignore")
    vendasI_df = vendasI_df.rename(columns={"Código da Proposta_y":"Código da Proposta","Cliente_y":"Cliente"})

    #Exportação
    output_path = os.path.join(caminho, 'RelatórioMesclado.xlsx')
    vendasI_df.to_excel(output_path, index=False)

    # -=-=-=-=- OPENPYXL -=-=-=-=-=-

    wb = load_workbook(output_path)
    ws = wb.active

    # Estilo de borda (fina)
    thin = Side(border_style="thin", color="000000")

    # Estilizar cabeçalho
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Ajustar largura das colunas
    for col in ws.columns:
        max_largura = 0
        col_letra = col[0].column_letter
        for cell in col:
            if cell.value:
                max_largura = max(max_largura, len(str(cell.value)))
            # Adicionar borda em todas as células
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # A partir da coluna de índice 4 (D pra frente) → centralizar
            if cell.col_idx >= 4 and cell.row > 1:  
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions[col_letra].width = max_largura + 2

    # Salvar
    wb.save(output_path)

# -=-=-=-=- TKINTER -=-=-=-=-=-

janela = ctk.CTk()

#Selecionador de Arquivo - POSVENDA
def file_arquivoPosVenda():
    global path_posVenda
    path_posVenda = ctk.filedialog.askopenfilename()
    print("Arquivo selecionado: ", path_posVenda)

#Selecionador de Arquivo - VENDASI
def file_arquivoVendasI():
    global path_vendasI
    path_vendasI = ctk.filedialog.askopenfilename()
    print("Arquivo selecionado: ", path_vendasI)

#Botão de Mesclar (Ativador da "def mesclarVendasI")
def button_mesclar():
    if path_posVenda and path_vendasI:
        mesclarVendasI(path_posVenda, path_vendasI)
        print("Relatório Mesclado")
    else:
        print('ERRO')


#Aparência da Janela
ctk.set_appearance_mode('dark')
janela.title("Mesclagem de Arquivos")
janela.geometry("500x300")  # largura x altura
janela.grid_columnconfigure(0, weight=1)  # coluna 0 ocupa espaço
janela.grid_columnconfigure(1, weight=2)  # coluna 1 maior

#Campos da VendasI 
text_arquivoPosVenda = ctk.CTkLabel(
    janela, text="📂 Escolha o arquivo do PosVenda:", anchor="w",
    font=("Arial", 15)
)
text_arquivoPosVenda.grid(row=0, column=0, padx=10, pady=18, sticky="w")

button_arquivoPosVenda = ctk.CTkButton(
    janela, text="Selecionar", command=file_arquivoPosVenda, corner_radius=15,
    font=("Arial", 15)
)
button_arquivoPosVenda.grid(row=0, column=1, padx=10, pady=18, sticky="ew")

#Campos da PosVenda
text_arquivoVendasI = ctk.CTkLabel(
    janela, text="📑 Escolha o arquivo do VendasI:", anchor="w",
    font=("Arial", 15)
)
text_arquivoVendasI.grid(row=1, column=0, padx=10, pady=18, sticky="w")

button_arquivoVendasI = ctk.CTkButton(
    janela, text="Selecionar", command=file_arquivoVendasI, corner_radius=15,
    font=("Arial", 15)
)
button_arquivoVendasI.grid(row=1, column=1, padx=10, pady=18, sticky="ew")

#Botão Mesclar
text_mesclarVendasI = ctk.CTkLabel(
    janela, text="🔄 Mesclar arquivos", font=("Arial", 15, "bold"),
)
text_mesclarVendasI.grid(row=2, column=0, columnspan=2, pady=(20, 5))

button_mesclarVendasI = ctk.CTkButton(
    janela,
    text="▶️ Iniciar Mesclagem",
    command=button_mesclar,
    fg_color="#4CAF50",
    hover_color="#45a049",
    height=40,
    corner_radius=20,
    font=("Arial", 18)
)
button_mesclarVendasI.grid(row=3, column=0, columnspan=2, padx=40, pady=15, sticky="ew")

janela.mainloop()

