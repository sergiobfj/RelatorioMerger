# merger/formatter.py

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

def formatar_excel(path_excel):
    wb = load_workbook(path_excel)
    ws = wb.active

    thin = Side(border_style="thin", color="000000")

    # Cabeçalho
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Colunas
    for col in ws.columns:
        max_largura = 0
        letra = col[0].column_letter

        for cell in col:
            if cell.value:
                max_largura = max(max_largura, len(str(cell.value)))

            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            if cell.col_idx >= 4 and cell.row > 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions[letra].width = max_largura + 2

    wb.save(path_excel)
